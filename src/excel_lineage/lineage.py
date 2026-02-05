from __future__ import annotations

from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import re


CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[^'!]+)!)?"
    r"(?P<cell>\$?[A-Za-z]{1,3}\$?\d+"
    r"(?:\s*:\s*\$?[A-Za-z]{1,3}\$?\d+)?)"
)


@dataclass
class Reference:
    sheet: str
    ref: str


@dataclass
class MetadataContext:
    top_headers: list[str] = field(default_factory=list)
    left_headers: list[str] = field(default_factory=list)


@dataclass
class MetricLineage:
    name: str
    sheet: str
    target: str
    formula: str | None
    references: list[Reference]
    metadata: MetadataContext


def _normalize_sheet(sheet: str | None, fallback: str) -> str:
    if not sheet:
        return fallback
    return sheet.strip("'")


def extract_references(formula: str, sheet: str) -> list[Reference]:
    references: list[Reference] = []
    for match in CELL_REF_RE.finditer(formula):
        ref_sheet = _normalize_sheet(match.group("sheet"), sheet)
        references.append(Reference(sheet=ref_sheet, ref=match.group("cell").replace(" ", "")))
    return references


def _extract_metadata(ws, cell_range: str) -> MetadataContext:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    top_headers: list[str] = []
    left_headers: list[str] = []
    if min_row > 1:
        for col in range(min_col, max_col + 1):
            value = ws.cell(row=min_row - 1, column=col).value
            if value is not None and str(value).strip():
                top_headers.append(str(value))
    if min_col > 1:
        for row in range(min_row, max_row + 1):
            value = ws.cell(row=row, column=min_col - 1).value
            if value is not None and str(value).strip():
                left_headers.append(str(value))
    return MetadataContext(top_headers=top_headers, left_headers=left_headers)


def _iter_named_destinations(wb, defined_name) -> Iterable[tuple[str, str]]:
    for sheet_name, cell_range in defined_name.destinations:
        if sheet_name is None:
            continue
        yield sheet_name, cell_range


def extract_lineage(path: str) -> list[MetricLineage]:
    wb = load_workbook(path, data_only=False)
    lineages: list[MetricLineage] = []
    for defined_name in wb.defined_names.definedName:
        if defined_name.is_external:
            continue
        if not defined_name.name:
            continue
        for sheet_name, cell_range in _iter_named_destinations(wb, defined_name):
            ws = wb[sheet_name]
            formula = None
            references: list[Reference] = []
            if cell_range and "!" not in cell_range:
                if ":" in cell_range:
                    min_col, min_row, _, _ = range_boundaries(cell_range)
                    cell = ws.cell(row=min_row, column=min_col)
                else:
                    cell = ws[cell_range]
                if cell.data_type == "f":
                    formula = f"={cell.value}"
                    references = extract_references(formula, sheet_name)
            metadata = _extract_metadata(ws, cell_range)
            lineages.append(
                MetricLineage(
                    name=defined_name.name,
                    sheet=sheet_name,
                    target=cell_range,
                    formula=formula,
                    references=references,
                    metadata=metadata,
                )
            )
    return lineages
